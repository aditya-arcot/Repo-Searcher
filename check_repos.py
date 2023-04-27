''' checks ADO repos for mentions of tables marked for deletion '''

import os
import sys
import time
import json
import logging
import re
import warnings

import openpyxl     # xlsm, xlsx
import xlrd         # for xls
import chardet
import git
from selenium import webdriver
from bs4 import BeautifulSoup



# use batch file to run, create log

# logged errors
    # no upstream branch
        # empty repo
        # skips pull
    # max retry for clone / pull
    # decoding failure
    # file not found
        # path too long (>260)



'''
testing

PRICE_VALUE - repos\FRR-Common
FORMULA_COMPONENT - repos\FRR-Common
HS_POSITION_SET - repos\InfohubPositionOutputService
'''
def test():
    path = 'repos\FRR-Common'
    write_repo_start(path, 1)
    write_repo_details(search_repo(path))



ADO_URL = 'https://dev.azure.com/bp-vsts/NAGPCCR/_apis/git/repositories?api-version=7.0'
MAX_GIT_ATTEMPTS = 5
RE_PATTERN_START = "(?<![a-z0-9_])"
RE_PATTERN_END = "(?![a-z0-9_])"
MAX_LINE_PREVIEW_LENGTH = 500

REPOS_FOLDER = 'repos'
REPOS_JSON = 'repos.json'
TABLE_NAMES_FILE = 'table_names.txt'
EXCLUDED_REPOS_FILE = 'excluded_repos.txt'
EXCLUDED_ENDINGS_FILE = 'excluded_endings.txt'
LAST_UPDATED_FILE = 'last_updated.txt'
FULL_RESULTS_FILE = 'results.txt'
SUMMARY_RESULTS_FILE = 'results_summary.txt'
AFFECTED_TABLE_NAMES_FILE = 'affected.txt'

table_names = []
affected_table_names = set()
excluded_repos = []
excluded_dirs = [".git", "obj", "bin", "service references", "connected services"]
excluded_endings = []
#logged_endings = ['.xls', '.xlsx', '.xlsm']
last_update = {}
encodings = ['default', 'utf-8', 'utf-16']



def search_repos():
    ''' checks repos for occurrences of table names '''

    with open(REPOS_JSON, 'r', encoding="utf-8") as json_file:
        repos_json = json.loads(''.join(json_file.readlines()))

        n_repos = repos_json['count']
        for count, i in enumerate(repos_json['value']):
            repo_name = i['name'].replace(' ', '%20') # no spaces in repo name
            repo_url = i['remoteUrl']
            repo_path = os.path.join(REPOS_FOLDER, repo_name)

            logger.info('repo %d/%d - %s - %s', count+1, n_repos, repo_name, repo_url)
            write_repo_start(repo_name, count+1)

            if repo_name in excluded_repos:
                logger.info('excluded from validation')
                write_repo_skipped()
                write_repo_end()
                continue

            if update_repo(repo_name, repo_path, repo_url):
                write_repo_details(search_repo(repo_path))

            else:
                logger.info('skipping search')
                write_repo_skipped()

            write_repo_end()



def update_repo(name, path, url=None):
    ''' updates repo local files '''

    if os.path.exists(path):
        logger.info('repo path already exists')

        try:
            git.Repo(path).git.rev_parse('@{upstream}') # look for upstream
        except git.GitCommandError:
            logger.error('no upstream branch - empty repo?')
            return False

        if (name not in last_update) or (time.time() - last_update[name] > 86400): # 1 day
            return update_repo_core(name, path, 'pull')

        logger.info('last update less than 1 day ago - skipping pull')
        return True

    logger.info('repo path does not exist')
    return update_repo_core(name, path, 'clone', url)



def update_repo_core(name, path, mode, url=None):
    ''' main logic for repo local file update '''

    attempts = 0

    while attempts < MAX_GIT_ATTEMPTS:
        if mode == 'pull':
            if attempt_git_command(path, 'pull'):
                break
        else:
            if attempt_git_command(path, 'clone', url):
                break
        attempts += 1
        logger.info('trying again in 2 seconds')
        time.sleep(2)

    if attempts == MAX_GIT_ATTEMPTS:
        if mode == 'pull':
            logger.error('max retries for repo pull - skipping')
        else:
            logger.error('max retries for repo clone - skipping')
        return False

    last_update[name] = time.time()
    return True



def attempt_git_command(path, mode, url=None):
    ''' attempts clone or pull and returns status '''

    try:
        if mode == 'clone':
            git.Repo.clone_from(url, path)
            logger.info('clone successful')
        else:
            logger.info(git.Repo(path).git.pull())
            logger.info('pull successful')
        return True

    except git.GitCommandError as err:
        if mode == 'clone':
            logger.error('clone failed - %s', err)
        else:
            logger.error('pull failed - %s', err)
        return False



def search_repo(repo_folder):
    ''' checks repo files for occurrences of table names '''

    matches = {}
    errors = []
    searched_subfolders = []
    searched_files = []

    for root, dirs, files in os.walk(repo_folder):
        dirs[:] = [_dir for _dir in dirs if _dir.lower() not in excluded_dirs]

        # only remove if .sln file exists
        if 'packages' in dirs:
            if len([file for file in files if file.endswith('.sln')]) != 0:
                dirs.remove('packages')

        searched_subfolders += [os.path.join(root, dir) for dir in dirs]

        files[:] = [file for file in files if not file.lower().endswith(tuple(excluded_endings))]

        for file in files:
            path = os.path.join(root, file)
            logger.info(path)

            if path.endswith('.xls'):
                search_legacy_spreadsheet_file(path, matches)
                searched_files.append(path)
            elif path.endswith(tuple(['.xlsm', '.xlsx'])):
                search_spreadsheet_file(path, matches)
                searched_files.append(path)
            else:
                if search_file(path, matches):
                    searched_files.append(path)
                else:
                    errors.append(path)

    return (errors, matches, searched_subfolders, searched_files)



def search_legacy_spreadsheet_file(path, matches):
    workbook = xlrd.open_workbook(path)

    for table_name in table_names:
        pattern = RE_PATTERN_START + table_name + RE_PATTERN_END

        for sheet in workbook.sheets():
            for n_row in range(sheet.nrows):
                for n_col in range(sheet.ncols):
                    cell = sheet.cell(n_row, n_col)
                    search_cell(cell, pattern, sheet.name, n_row, n_col, matches, path, table_name)



def search_spreadsheet_file(path, matches):
    warnings.simplefilter(action='ignore', category=UserWarning)

    workbook = openpyxl.load_workbook(path, read_only=True)

    for table_name in table_names:
        pattern = RE_PATTERN_START + table_name + RE_PATTERN_END

        for sheet in workbook.worksheets:
            for n_row, row in enumerate(sheet.iter_rows()):
                for n_col, cell in enumerate(row):
                    search_cell(cell, pattern, sheet.title, n_row, n_col, matches, path, table_name)

    warnings.resetwarnings()



def search_cell(cell, pattern, sheet_name, n_row, n_col, matches, path, table_name):
    if not cell.value:
        return

    val = str(cell.value).lower()

    if re.search(pattern, val):
        if len(val) > MAX_LINE_PREVIEW_LENGTH:
            val = 'VALUE TOO LONG, LOOK AT FILE'

        log_string = f'sheet {sheet_name} row {n_row} col {n_col} - {val}'

        logger.info('match - %s - %s', table_name, log_string)
        add_new_match(matches, path, table_name, log_string)



def search_file(path, matches):
    ''' checks file for occurrences of table names '''

    decoding_result = decode_file(path)
    if not decoding_result[0]:
        return False

    lines = decoding_result[1]

    for table_name in table_names:
        pattern = RE_PATTERN_START + table_name + RE_PATTERN_END

        for line_num, line in enumerate(lines):
            line = line.lower().strip()
            if re.search(pattern, line):
                if len(line) > MAX_LINE_PREVIEW_LENGTH:
                    line = 'LINE TOO LONG, LOOK AT FILE'

                log_string = f'line {line_num} - {line}'

                logger.info('match - %s - %s', table_name, log_string)
                add_new_match(matches, path, table_name, log_string)

    return True



def add_new_match(matches, path, table_name, note):
    if not path in matches:
        matches[path] = {}

    if not table_name in matches[path]:
        matches[path][table_name] = []

    matches[path][table_name].append(note)

    affected_table_names.add(table_name)



def decode_file(path):
    ''' decodes file for reading '''

    for encoding in encodings:
        try:
            if encoding == 'default':
                with open(path, 'r') as cur_file: # intentionally didn't specify encoding
                    lines = [line.lower() for line in cur_file.readlines()]
            else:
                with open(path, 'r', encoding=encoding) as cur_file:
                    lines = [line.lower() for line in cur_file.readlines()]

            logger.info('decoding success using %s', encoding)
            return (True, lines)

        except FileNotFoundError:
            logger.error('file not found')
            return (False,)

        except (UnicodeDecodeError, UnicodeError):
            pass


    with open(path, 'rb') as binary_file:
        data = binary_file.read()

    encoding = chardet.detect(data)['encoding']

    if encoding is None:
        logger.error('decoding failure type 1')
        return (False,)

    try:
        with open(path, 'r', encoding=encoding) as cur_file:
            lines = [line.lower() for line in cur_file.readlines()]

        logger.info('decoding success using %s', encoding)
        return (True, lines)

    except (UnicodeDecodeError, UnicodeError):
        logger.error('decoding failure type 2')
        return (False,)



def create_repos_json():
    ''' gets repos list using API, persists in json '''

    driver = webdriver.Chrome()
    driver.get(ADO_URL)

    input() # press enter once loaded

    body = BeautifulSoup(driver.page_source, 'html.parser').body
    driver.close()
    for pre_tag in body.find_all('pre'):
        pre_tag.unwrap()
    body_content = body.contents

    with open(REPOS_JSON, 'w', encoding="utf-8") as json_file:
        json_file.write(body_content[0])



def read_table_names():
    ''' reads table names for validation '''

    with open(TABLE_NAMES_FILE, 'r', encoding="utf-8") as tables:
        for line in tables.readlines():
            line = line.strip()

            if not line == "":
                table_names.append(line.lower())



def read_excluded_repo_names():
    ''' reads excluded repos to skip validation '''

    with open(EXCLUDED_REPOS_FILE, 'r', encoding="utf-8") as excluded_file:
        for line in excluded_file.readlines():
            line = line.strip()

            if (not line.startswith("#")) and (not line == ""):
                excluded_repos.append(line.replace(' ', '%20')) # no spaces in repo name



def read_excluded_endings():
    ''' reads excluded file endings from validation '''

    with open(EXCLUDED_ENDINGS_FILE, 'r', encoding='utf-8') as excluded_ext_file:
        for line in excluded_ext_file.readlines():
            line = line.strip().lower()

            if (not line.startswith('#')) and (not line == ""):
                excluded_endings.append(line)



def read_last_updated_info():
    ''' reads last updated info for repos '''

    with open(LAST_UPDATED_FILE, 'r', encoding="utf-8") as update_file:
        for line in update_file.readlines():
            split = line.strip().split('\t')
            last_update[split[0]] = float(split[1])



def write_last_updated_info():
    ''' writes last updated info for repos '''

    with open(LAST_UPDATED_FILE, 'w', encoding="utf-8") as update_file:
        for repo, update_time in last_update.items():
            update_file.write(repo + '\t' + str(update_time) + '\n')



def init_results_files():
    ''' creates empty results files '''

    with open(FULL_RESULTS_FILE, 'w', encoding='utf-8') as file:
        file.write('')
    with open(SUMMARY_RESULTS_FILE, 'w', encoding='utf-8') as file:
        file.write('')



def write_full_results_line(line):
    ''' writes line to full results file '''

    with open(FULL_RESULTS_FILE, 'a', encoding='utf-8') as file:
        file.write(line + '\n')



def write_summary_results_line(line):
    ''' writes line to summary results file '''

    with open(SUMMARY_RESULTS_FILE, 'a', encoding='utf-8') as file:
        file.write(line + '\n')



def write_repo_start(name, num):
    ''' writes repo start info to results files '''

    write_full_results_line(name + f' ({num})')
    write_summary_results_line(name + f' ({num})')



def write_repo_skipped():
    ''' writes repo skipped to results files '''

    write_full_results_line('\tskipped')
    write_summary_results_line('\tskipped')



def write_repo_details(details):
    ''' writes repo validation details to results files '''

    errors, matches, searched_subfolders, searched_files = details

    if len(matches) > 0:
        count = 0

        write_full_results_line('\tmatches:')
        for path in matches.keys():
            write_full_results_line(f'\t\t{path}')
            for table_name in matches[path].keys():
                write_full_results_line(f'\t\t\t{table_name}')
                for match in matches[path][table_name]:
                    write_full_results_line(f'\t\t\t\t{match}')
                    count += 1

        write_summary_results_line(f'\tmatches: {count}')

    if len(errors) > 0:
        write_full_results_line('\terrors:')
        for error in errors:
            write_full_results_line('\t\t' + error)

        write_summary_results_line(f'\terrors: {len(errors)}')

    if len(searched_subfolders) > 0:
        write_full_results_line('\tsubfolders searched:')
        for folder in searched_subfolders:
            write_full_results_line('\t\t' + folder)

        write_summary_results_line(f'\tsubfolders searched: {len(searched_subfolders)}')

    if len(searched_files) > 0:
        write_full_results_line('\tfiles searched:')
        for file in searched_files:
            write_full_results_line('\t\t' + file)

        write_summary_results_line(f'\tfiles searched: {len(searched_files)}')



def write_repo_end():
    ''' writes repo end to results files '''

    write_full_results_line('\n\n')
    write_summary_results_line('\n\n')



def write_affected_tables():
    sorted_affected_table_names = sorted(affected_table_names)
    with open(AFFECTED_TABLE_NAMES_FILE, 'w', encoding='utf-8') as file:
        for name in sorted_affected_table_names:
            file.write(name + '\n')



def main():
    ''' driver for validation process '''

    if not os.path.isfile(REPOS_JSON):
        create_repos_json()

    if not os.path.isfile(TABLE_NAMES_FILE):
        logger.critical('table names not given')
        sys.exit()
    read_table_names()

    if os.path.isfile(EXCLUDED_REPOS_FILE):
        read_excluded_repo_names()

    if os.path.isfile(EXCLUDED_ENDINGS_FILE):
        read_excluded_endings()

    if os.path.isfile(LAST_UPDATED_FILE):
        read_last_updated_info()

    init_results_files()
    search_repos()
    write_last_updated_info()
    write_affected_tables()



if __name__ == "__main__":
    logging.basicConfig(
        level=logging.DEBUG,
        format='%(asctime)s - LOGGING.%(levelname)s - %(message)s'
    )
    logger = logging.getLogger(__name__)

    main()
