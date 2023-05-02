import openpyxl
import os
import re

RE_PATTERN_START = "(?<![a-z0-9_])"
RE_PATTERN_END = "(?![a-z0-9_])"

path = "ODS_DEPENDECY_DETAILS.xlsx"
workbook = openpyxl.load_workbook(path)

with open("table_names.txt") as file:
    names = [line.strip() for line in file.readlines()]

def search_workbook(name, workbook):
    matches = []
    pattern = RE_PATTERN_START + name + RE_PATTERN_END

    for sheet in workbook.worksheets:
        if sheet.title == 'Objects to Validate':
            continue

        for row in sheet.iter_rows():
            for n_col, cell in enumerate(row):
                val = str(cell.value)

                if re.search(pattern, val):
                    col_header = sheet.cell(row=1, column=n_col+1).value.lower() #1-indexed

                    matches.append([sheet.title, col_header])

    assert len(matches) > 0
    write_table(name, matches)

def write_table(name, matches):
    with open('table_context.txt', 'a') as f:
        f.write(f'{name}\t' + ', '.join([sheet + ' (' + col + ')' for sheet, col in matches]) + '\n')

    with open('schema_table_pairs.txt', 'a') as f:
        for sheet, _ in matches:
            f.write(f'{sheet}\t{name}\n')

with open('schema_table_pairs.txt', 'w') as f:
    f.write('')

with open('table_context.txt', 'w') as f:
    f.write('')

for name in names:
    search_workbook(name, workbook)


"""
table_name
functions
synonyms
views
child tables
parent tables
packages
stored procedures
"""


