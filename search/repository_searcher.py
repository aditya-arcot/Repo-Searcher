''' contains repository searcher class '''

import os
import re
import json
import requests
import xlrd
import openpyxl
from repository import ADORepository
from logging_manager import LoggingManager
from configuration_manager import ConfigurationManager
from constants import ConfigEnum, RepoSearchModeEnum, Constants
from input_manager import InputManager
from results_writer import ResultsWriter, RepoSearchResults

class RepositorySearcher:
    ''' used for searching repositories for specific words '''
    max_line_preview_length = 1000
    search_pattern = '(?<![a-z0-9_]){word}(?![a-z0-9_])'
    line_too_long_msg = 'VALUE TOO LONG, LOOK AT FILE'

    def __init__(self, _logger:LoggingManager) -> None:
        self.logger = _logger
        self.__init_input()
        self.__init_config()
        self.__init_writer()
        self.__create_repos_json()
        self.__found_words = set()

    def __init_input(self) -> None:
        ''' initialize input manager, set options '''
        self.input_manager = InputManager(self.logger)
        self.repo_mode = self.input_manager.get_repo_mode()
        self.search_excel = self.input_manager.get_search_excel()

    def __init_config(self) -> None:
        ''' initialize configuration manager, set config info '''
        self.config = ConfigurationManager(self.logger)
        self.config.add_pat()
        self.config.add_search_words()
        self.config.add_excluded_files()
        self.config.add_excluded_folders()
        self.config.add_last_updated_info()
        if self.repo_mode == RepoSearchModeEnum.INCLUDE_MODE.value:
            self.config.add_included_repo_names()
        else:
            self.config.add_excluded_repo_names()

    def __init_writer(self) -> None:
        ''' initialize results writer, write config section '''
        self.writer = ResultsWriter(self.logger)
        self.writer.write_config(RepoSearchModeEnum(self.repo_mode).name, self.config)

    def __create_repos_json(self) -> None:
        '''
        create json with repository information from ADO API
        see learn.microsoft.com/en-us/rest/api/azure/devops/git/?view=azure-devops-rest-7.0
        '''
        git_repos_url = 'https://dev.azure.com/bp-vsts/NAGPCCR/_apis/git/' + \
                            'repositories?api-version=7.0'
        resp = requests.get(git_repos_url, \
                            auth=('user', self.config.get(ConfigEnum.PAT.name)), \
                            timeout=10)

        if resp.status_code != 200:
            self.logger.critical(f'error requesting repos - {resp.status_code}')

        try:
            with open(ConfigEnum.REPOS.value, 'w', encoding='utf-8') as json_file:
                json.dump(resp.json(), json_file, indent=4)
        except requests.exceptions.JSONDecodeError:
            self.logger.critical('error parsing repos json info')

    def search(self) -> None:
        ''' search through repositories for specified words '''
        with open(ConfigEnum.REPOS.value, 'r', encoding='utf-8') as json_file:
            data = json.loads(''.join(json_file.readlines()))
            n_repos = data['count']
            self.logger.info(f"{n_repos} repos")

            for num, repo_data in enumerate(data['value']):
                repo = self.__create_repo_obj(num, repo_data, n_repos)
                if repo is None:
                    continue

                if repo.update(): # update unnecessary or successful
                    details = self.__search_repo(repo)
                    self.writer.write_repo_results(details)
                else:
                    self.__skip_repo('update unsuccessful')

                self.writer.write_repo_end()

        self.writer.write_found_words(self.__found_words)
        self.__write_last_update()

    def __check_included_repo(self, repo_name:str) -> bool:
        ''' check if repo is included in search '''
        if self.repo_mode == RepoSearchModeEnum.INCLUDE_MODE.value:
            return repo_name in self.config.get(ConfigEnum.INCLUDED_REPOS.name)
        return repo_name not in self.config.get(ConfigEnum.EXCLUDED_REPOS.name)

    def __skip_repo(self, msg:str) -> None:
        ''' log, write info for skipped repo '''
        self.logger.info(msg)
        self.writer.write_repo_skip(msg)
        self.writer.write_repo_end()

    def __create_repo_obj(self, num:int, repo_data:dict, n_repos:int) -> ADORepository:
        ''' returns ADO repository object '''
        try:
            name = repo_data['name'].lower().replace(' ', Constants.ENCODED_SPACE)
            url = repo_data['remoteUrl']
            _id = repo_data['id']
        except KeyError:
            self.logger.critical('issue parsing repos json')

        self.logger.info(f'repo {num+1}/{n_repos} - {name}')

        self.writer.write_repo_start(name)
        if not self.__check_included_repo(name):
            self.__skip_repo('not included in search')
            return None
        self.logger.info('included in search')

        default_branch = 'NONE'
        if 'defaultBranch' in repo_data:
            default_branch = repo_data['defaultBranch']

        repo = ADORepository(self.logger, self.config, name, url, _id, \
                             os.path.join('repos', name), default_branch)
        return repo

    def __search_repo(self, repo:ADORepository) -> RepoSearchResults:
        ''' search local repository files for search words '''
        search_results = RepoSearchResults()

        for root, dirs, files in os.walk(repo.path):
            excluded_folders = self.config.get(ConfigEnum.EXCLUDED_FOLDERS.name)
            excluded_files = self.config.get(ConfigEnum.EXCLUDED_FILES.name)

            skipped_dirs = [_dir for _dir in dirs if _dir.lower() in excluded_folders]
            skipped_files = [file for file in files if \
                             file.lower().endswith(tuple(excluded_files))]

            search_results.skipped_folders += [os.path.join(root, _dir) + os.sep \
                                               for _dir in skipped_dirs]
            search_results.skipped_files += [os.path.join(root, file) for file in skipped_files]

            # modify in place to prune search tree
            dirs[:] = [_dir for _dir in dirs if _dir not in skipped_dirs]
            files[:] = [file for file in files if file not in skipped_files ]

            search_results.folders += [os.path.join(root, dir) for dir in dirs]

            for file in files:
                path = os.path.join(root, file)
                logger.info(f'file - {path}', stdout=False)
                self.__search_file(path, search_results)

        return search_results

    def __search_file(self, path, search_results:RepoSearchResults) -> None:
        ''' searches file for specified words '''
        # Excel files
        if path.endswith(tuple(['.xls', '.xlsm', '.xlsx'])):
            if not self.search_excel:
                search_results.skipped_files.append(path)
            else:
                search_results.files.append(path)
                if path.endswith('.xls'):
                    self.__search_legacy_spreadsheet_file(path, search_results.matches)
                else:
                    self.__search_spreadsheet_file(path, search_results.matches)
        else:
            self.__search_ordinary_file(path, search_results)

    def __search_legacy_spreadsheet_file(self, path:str, matches:dict) -> None:
        ''' searches .xls file by iterating over words, sheets, cells '''
        workbook = xlrd.open_workbook(path)
        for search_word in self.config.get(ConfigEnum.SEARCH_WORDS.name):
            pattern = self.search_pattern.format(word=search_word)
            for sheet in workbook.sheets():
                for n_row in range(sheet.nrows):
                    for n_col in range(sheet.ncols):
                        cell = sheet.cell(n_row, n_col)
                        self.__search_cell(cell, sheet.name, n_row, n_col, matches, \
                                           path, pattern, search_word)

    def __search_spreadsheet_file(self, path:str, matches:dict) -> None:
        ''' searches .xlsm or .xlsx file by iterating over words, sheets, cells'''
        workbook = openpyxl.load_workbook(path, read_only=True)
        for search_word in self.config.get(ConfigEnum.SEARCH_WORDS.name):
            pattern = self.search_pattern.format(word=search_word)
            for sheet in workbook.worksheets:
                for n_row, row in enumerate(sheet.iter_rows()):
                    for n_col, cell in enumerate(row):
                        self.__search_cell(cell, sheet.title, n_row, n_col, matches, \
                                           path, pattern, search_word)

    def __search_cell(self, cell, sheet_name:str, n_row:int, n_col:int,
                      matches:dict, path:str, pattern:str, search_word:str) -> None:
        ''' searches cell for search pattern '''
        if not cell.value:
            return

        val = str(cell.value).lower()

        if re.search(pattern, val):
            if len(val) > self.max_line_preview_length:
                val = self.line_too_long_msg
            log_string = f'sheet {sheet_name}, row {n_row}, col {n_col} - {val}'
            logger.info(f'match - {search_word} - {log_string}', stdout=False)
            self.__add_new_match(matches, path, search_word, log_string)

    def __search_ordinary_file(self, path:str, search_results:RepoSearchResults) -> None:
        ''' searches non-spreadsheet file for specified words '''
        lines = self.__decode_file(path, search_results)
        if len(lines) == 0:
            return

        for search_word in self.config.get(ConfigEnum.SEARCH_WORDS.name):
            pattern = self.search_pattern.format(word=search_word)
            for line_num, line in enumerate(lines):
                line = line.lower().strip()

                if re.search(pattern, line):
                    if len(line) > self.max_line_preview_length:
                        line = self.line_too_long_msg
                    log_string = f'line {line_num + 1} - {line}'
                    logger.info(f'match - {search_word} - {log_string}', stdout=False)
                    self.__add_new_match(search_results.matches, path, search_word, log_string)

    def __decode_file(self, path:str, search_results:RepoSearchResults) -> list:
        ''' attempts decoding file, returns lines '''        
        try:
            with open(path, 'r', encoding='utf-8', errors='ignore') as cur_file:
                lines = [line.lower() for line in cur_file.readlines()]
            logger.info('decoding success', stdout=False)
            search_results.files.append(path)
            return lines

        except FileNotFoundError:
            logger.error(f'file not found - path too long - {path}', stdout=False)

        except (UnicodeDecodeError, UnicodeError):
            logger.error('decoding failure', stdout=False)

        search_results.errors.append(path)
        return []

    def __add_new_match(self, matches:dict, path:str, search_word:str, note:str) -> None:
        ''' adds new match to matches dict '''
        if not path in matches:
            matches[path] = {}

        if not search_word in matches[path]:
            matches[path][search_word] = []

        matches[path][search_word].append(note)

        self.__found_words.add(search_word)

    def __write_last_update(self) -> None:
        ''' writes last updated info to file '''
        with open(ConfigEnum.LAST_UPDATE.value, 'w', encoding="utf-8") as update_file:
            for repo, update_time in self.config.get(ConfigEnum.LAST_UPDATE.name).items():
                update_file.write(repo + Constants.TAB + str(update_time) + Constants.NEWLINE)


if __name__ == '__main__':
    logger = LoggingManager()
    repoSearcher = RepositorySearcher(logger)
    repoSearcher.search()
