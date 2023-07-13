''' searches ADO repos for occurrences of given words '''

import os
import sys
import time
import json
import logging
import re
import warnings
import datetime

import openpyxl     # xlsm, xlsx
import xlrd         # xls
import chardet
import git
from selenium import webdriver
from bs4 import BeautifulSoup


## TODO
# consider regex options
# user options


"""
handled errors

no upstream branch
    empty repo
    resolution - skip repo
max retry for clone / pull
    resolution - skip repo
file not found
    path too long (>260)
    resolution - skip file
decoding failure
    resolution - skip file
"""



LOGS_FOLDER = 'logs'

REPOS_FOLDER = 'repos'
REPOS_JSON = 'repos_info.json'
LAST_UPDATED_FILENAME = 'repos_last_update.txt'

INPUT_FOLDER = 'input'
SEARCH_WORDS_FILENAME = 'words.txt'
INCLUDED_REPOS_FILENAME = 'included_repos.txt'
EXCLUDED_REPOS_FILENAME = 'excluded_repos.txt'
EXCLUDED_ENDINGS_FILENAME = 'excluded_endings.txt'
EXCLUDED_FOLDERS_FILENAME = 'excluded_folders.txt'

OUTPUT_FOLDER = 'output'
FULL_RESULTS_FILENAME = 'results.txt'
FULL_RESULTS_FILE = ''
SUMMARY_RESULTS_FILENAME = 'results_summary.txt'
SUMMARY_RESULTS_FILE = ''
FOUND_WORDS_FILENAME = 'found_words.txt'

ADO_URL = 'https://dev.azure.com/bp-vsts/NAGPCCR/_apis/git/repositories?api-version=7.0'
MAX_GIT_ATTEMPTS = 5
RE_PATTERN_START = "(?<![a-z0-9_])"
RE_PATTERN_END = "(?![a-z0-9_])"
MAX_LINE_PREVIEW_LENGTH = 1000
DAY_IN_SECONDS = 86400

INCLUDE_MODE, EXCLUDE_MODE = range(2)
REPO_MODE = 0

search_words = []
found_words = set()
included_repos = []
excluded_repos = []
excluded_folders = []
excluded_endings = []
last_update = {}
encodings = ['default', 'utf-8', 'utf-16']



def search_repos():
    ''' searches repos for occurrences of search words '''

    with open(REPOS_JSON, 'r', encoding="utf-8") as json_file:
        repos_json = json.loads(''.join(json_file.readlines()))

        n_repos = repos_json['count']
        for count, i in enumerate(repos_json['value']):
            repo_name = i['name'].lower().replace(' ', '%20') # no spaces in repo name
            repo_url = i['remoteUrl']
            repo_path = os.path.join(REPOS_FOLDER, repo_name)

            logger.info('repo %d/%d - %s - %s', count+1, n_repos, repo_name, repo_url)
            print(f'searching repo {count+1}/{n_repos} - {repo_name}')
            write_repo_start(repo_name, count+1)

            if ((REPO_MODE == INCLUDE_MODE) and (repo_name not in included_repos)) or \
                ((REPO_MODE == EXCLUDE_MODE) and (repo_name in excluded_repos)):
                logger.info('excluded from validation')
                write_repo_skipped()
                write_repo_end()
                continue

            if update_repo(repo_name, repo_path, repo_url): # update unnecessary or successful
                write_repo_details(search_repo(repo_path))
            else: # update failed
                logger.info('skipping search')
                write_repo_skipped()

            write_repo_end()



def search_repo(repo_folder):
    ''' checks repo files for occurrences of search words '''

    matches = {}
    errors = []
    searched_subfolders = []
    searched_files = []

    for root, dirs, files in os.walk(repo_folder):
        # removed excluded folders
        dirs[:] = [_dir for _dir in dirs if _dir.lower() not in excluded_folders]

        # only remove if .sln file exists
        if 'packages' in dirs:
            if len([file for file in files if file.endswith('.sln')]) != 0:
                dirs.remove('packages')

        searched_subfolders += [os.path.join(root, dir) for dir in dirs]

        files[:] = [file for file in files if not file.lower().endswith(tuple(excluded_endings))]

        for file in files:
            path = os.path.join(root, file)
            logger.info(path)

            if path.endswith('.xls'):
                search_legacy_spreadsheet_file(path, matches)
                searched_files.append(path)
            elif path.endswith(tuple(['.xlsm', '.xlsx'])):
                search_spreadsheet_file(path, matches)
                searched_files.append(path)
            else:
                if search_file(path, matches):
                    searched_files.append(path)
                else:
                    errors.append(path)

    return (errors, matches, searched_subfolders, searched_files)



def search_legacy_spreadsheet_file(path, matches):
    ''' searches xls file for search words '''

    workbook = xlrd.open_workbook(path)

    for search_word in search_words:
        pattern = RE_PATTERN_START + search_word + RE_PATTERN_END

        for sheet in workbook.sheets():
            for n_row in range(sheet.nrows):
                for n_col in range(sheet.ncols):
                    cell = sheet.cell(n_row, n_col)
                    search_cell(cell, pattern, sheet.name, n_row, n_col, matches, path, search_word)



def search_spreadsheet_file(path, matches):
    ''' searches xlsx/xlsm file for search words '''

    warnings.simplefilter(action='ignore', category=UserWarning)

    workbook = openpyxl.load_workbook(path, read_only=True)

    for search_word in search_words:
        pattern = RE_PATTERN_START + search_word + RE_PATTERN_END

        for sheet in workbook.worksheets:
            for n_row, row in enumerate(sheet.iter_rows()):
                for n_col, cell in enumerate(row):
                    search_cell(cell, pattern, sheet.title, n_row, n_col, \
                                matches, path, search_word)

    warnings.resetwarnings()



def search_cell(cell, pattern, sheet_name, n_row, n_col, matches, path, search_word):
    ''' searches spreadsheet cell for search word '''

    if not cell.value:
        return

    val = str(cell.value).lower()

    if re.search(pattern, val):
        if len(val) > MAX_LINE_PREVIEW_LENGTH:
            val = 'VALUE TOO LONG, LOOK AT FILE'

        log_string = f'sheet {sheet_name} row {n_row} col {n_col} - {val}'

        logger.info('match - %s - %s', search_word, log_string)
        add_new_match(matches, path, search_word, log_string)



def search_file(path, matches):
    ''' searches file for search words '''

    decoding_result = decode_file(path)
    if not decoding_result[0]:
        return False

    lines = decoding_result[1]

    for search_word in search_words:
        pattern = RE_PATTERN_START + search_word + RE_PATTERN_END

        for line_num, line in enumerate(lines):
            line = line.lower().strip()
            if re.search(pattern, line):
                if len(line) > MAX_LINE_PREVIEW_LENGTH:
                    line = 'LINE TOO LONG, LOOK AT FILE'

                log_string = f'line {line_num} - {line}'

                logger.info('match - %s - %s', search_word, log_string)
                add_new_match(matches, path, search_word, log_string)

    return True



def decode_file(path):
    ''' decodes file for reading '''

    for encoding in encodings:
        try:
            if encoding == 'default':
                with open(path, 'r') as cur_file: # intentionally didn't specify encoding
                    lines = [line.lower() for line in cur_file.readlines()]
            else:
                with open(path, 'r', encoding=encoding) as cur_file:
                    lines = [line.lower() for line in cur_file.readlines()]

            logger.info('decoding success using %s', encoding)
            return (True, lines)

        except FileNotFoundError:
            logger.error('file not found')
            return (False,)

        except (UnicodeDecodeError, UnicodeError):
            pass


    with open(path, 'rb') as binary_file:
        data = binary_file.read()

    encoding = chardet.detect(data)['encoding']

    if encoding is None:
        logger.error('decoding failure type 1')
        return (False,)

    try:
        with open(path, 'r', encoding=encoding) as cur_file:
            lines = [line.lower() for line in cur_file.readlines()]

        logger.info('decoding success using %s', encoding)
        return (True, lines)

    except (UnicodeDecodeError, UnicodeError):
        logger.error('decoding failure type 2')
        return (False,)



def add_new_match(matches, path, search_word, note):
    ''' adds new match to dictionary '''

    if not path in matches:
        matches[path] = {}

    if not search_word in matches[path]:
        matches[path][search_word] = []

    matches[path][search_word].append(note)

    found_words.add(search_word)



def update_repo(name, path, url=None):
    ''' updates repo local files '''

    if os.path.exists(path):
        logger.info('repo path already exists')

        try:
            git.Repo(path).git.rev_parse('@{upstream}') # look for upstream
        except git.GitCommandError:
            logger.error('no upstream branch - empty repo?')
            return False

        if (name not in last_update) or (time.time() - last_update[name] > DAY_IN_SECONDS):
            return update_repo_core(name, path, 'pull')

        logger.info('last update less than 1 day ago - skipping pull')
        return True

    logger.info('repo path does not exist')
    return update_repo_core(name, path, 'clone', url)



def update_repo_core(name, path, mode, url=None):
    ''' main logic for repo local files update '''

    attempts = 0

    while attempts < MAX_GIT_ATTEMPTS:
        if mode == 'pull':
            if attempt_git_command(path, 'pull'): # success
                break
        else:
            if attempt_git_command(path, 'clone', url): # success
                break
        attempts += 1
        logger.info('trying again in 2 seconds')
        time.sleep(2)

    if attempts == MAX_GIT_ATTEMPTS:
        if mode == 'pull':
            logger.error('max retries for repo pull - skipping')
        else:
            logger.error('max retries for repo clone - skipping')
        return False

    last_update[name] = time.time()
    return True



def attempt_git_command(path, mode, url=None):
    ''' attempts clone or pull and returns status '''

    try:
        if mode == 'clone':
            git.Repo.clone_from(url, path)
            logger.info('clone successful')
        else:
            logger.info(git.Repo(path).git.pull())
            logger.info('pull successful')
        return True

    except git.GitCommandError as err:
        logger.error('%s failed - %s', mode, err)
        return False



def check_repos_json():
    ''' checks for valid repos info json'''

    if not os.path.exists(REPOS_JSON):
        return False

    with open(REPOS_JSON, 'r', encoding="utf-8") as json_file:
        repos_json = json.loads(''.join(json_file.readlines()))

        try:
            dtm = float(repos_json['dtm'])
        except ValueError:
            logger.warning('json update dtm could not be converted to float')
            return False
        except KeyError:
            logger.error('json does not contain dtm')
            return False

        return time.time() - dtm < DAY_IN_SECONDS



def create_repos_json():
    ''' gets repo details, persists in json '''

    driver = webdriver.Chrome()
    driver.get(ADO_URL)

    input() # press enter once loaded

    body = BeautifulSoup(driver.page_source, 'html.parser').body
    driver.close()
    for pre_tag in body.find_all('pre'):
        pre_tag.unwrap()

    json_dict = json.loads(body.contents[0])
    json_dict["dtm"] = time.time()

    with open(REPOS_JSON, 'w', encoding="utf-8") as json_file:
        json.dump(json_dict, json_file, indent=4)



def set_repo_mode():
    ''' sets repo mode from user input '''
    global REPO_MODE
    if input('enter repo mode (include - i / I, exclude - anything else): ') in ('i', 'I'):
        REPO_MODE = INCLUDE_MODE
        return
    REPO_MODE = EXCLUDE_MODE



def read_input_file(filename, critical=False):
    ''' checks for input file, returns lines '''

    path = os.path.join(INPUT_FOLDER, filename)
    if not os.path.exists(path):
        msg = f'{filename} does not exist'
        if critical:
            logger.critical(msg)
            sys.exit()
        else:
            logger.warning(msg)
            return []

    out = []
    with open(path, 'r', encoding="utf-8") as file:
        for line in file.readlines():
            line = line.strip().lower()
            if (not line.startswith("#")) and (not line == ""):
                out.append(line)
    return out



def read_repo_names(filename):
    ''' returns formatted repo names '''
    return [name.replace(' ', '%20') for name in read_input_file(filename)]



def read_last_updated_info():
    ''' reads last updated info for repos '''

    if not os.path.exists(LAST_UPDATED_FILENAME):
        logger.warning('%s does not exist', LAST_UPDATED_FILENAME)
        return

    with open(LAST_UPDATED_FILENAME, 'r', encoding="utf-8") as update_file:
        for line in update_file.readlines():
            split = line.strip().split('\t')
            last_update[split[0]] = float(split[1])



def init_results_files():
    ''' initializes results files '''

    global FULL_RESULTS_FILE, SUMMARY_RESULTS_FILE

    FULL_RESULTS_FILE = open(os.path.join(OUTPUT_FOLDER, FULL_RESULTS_FILENAME), \
                             'w', encoding='utf-8')
    SUMMARY_RESULTS_FILE = open(os.path.join(OUTPUT_FOLDER, SUMMARY_RESULTS_FILENAME), \
                                'w', encoding='utf-8')

    lines = [
        '=== CONFIG ===',
        f'log file - {LOGFILE}',
        f'repo mode - {REPO_MODE}',
        format_config_section('included repos', included_repos) if REPO_MODE == INCLUDE_MODE else\
            format_config_section('excluded repos', excluded_repos),
        format_config_section('excluded endings', excluded_endings),
        format_config_section('excluded dirs', excluded_folders),
        format_config_section('search words', search_words) + '\n',
        '=== REPOS ==='
    ]
    for line in lines:
        write_full_results_line(line)

    SUMMARY_RESULTS_FILE.write('')



def format_config_section(start, lst):
    ''' formats config section into output string '''
    out = f'{start}'
    if len(lst) > 0:
        out += '\n\t' + '\n\t'.join(lst)
    return out



def write_last_updated_info():
    ''' writes last updated info for repos '''
    with open(LAST_UPDATED_FILENAME, 'w', encoding="utf-8") as update_file:
        for repo, update_time in last_update.items():
            update_file.write(repo + '\t' + str(update_time) + '\n')



def write_repo_start(name, num):
    ''' writes repo start info to results files '''
    write_full_results_line(name + f' ({num})')
    write_summary_results_line(name + f' ({num})')



def write_repo_skipped():
    ''' writes repo skipped to results files '''
    write_full_results_line('\tskipped')
    write_summary_results_line('\tskipped')



def write_repo_details(details):
    ''' writes repo validation details to results files '''

    errors, matches, searched_subfolders, searched_files = details

    if len(matches) > 0:
        count = 0

        write_full_results_line('\tmatches:')
        for path in matches.keys():
            write_full_results_line(f'\t\t{path}')
            for search_word in matches[path].keys():
                write_full_results_line(f'\t\t\t{search_word}')
                for match in matches[path][search_word]:
                    write_full_results_line(f'\t\t\t\t{match}')
                    count += 1

        write_summary_results_line(f'\tmatches: {count}')

    if len(errors) > 0:
        write_full_results_line('\terrors:')
        for error in errors:
            write_full_results_line('\t\t' + error)

        write_summary_results_line(f'\terrors: {len(errors)}')

    if len(searched_subfolders) > 0:
        write_full_results_line('\tsubfolders searched:')
        for folder in searched_subfolders:
            write_full_results_line('\t\t' + folder)

        write_summary_results_line(f'\tsubfolders searched: {len(searched_subfolders)}')

    if len(searched_files) > 0:
        write_full_results_line('\tfiles searched:')
        for file in searched_files:
            write_full_results_line('\t\t' + file)

        write_summary_results_line(f'\tfiles searched: {len(searched_files)}')



def write_repo_end():
    ''' writes repo end to results files '''
    write_full_results_line('\n\n')
    write_summary_results_line('\n\n')



def write_found_words():
    ''' writes found search words to file '''
    with open(os.path.join(OUTPUT_FOLDER, FOUND_WORDS_FILENAME), 'w', encoding='utf-8') as file:
        for name in sorted(found_words):
            file.write(name + '\n')



def write_full_results_line(line):
    ''' writes line to full results file '''
    FULL_RESULTS_FILE.write(line + '\n')
    FULL_RESULTS_FILE.flush()




def write_summary_results_line(line):
    ''' writes line to summary results file '''
    SUMMARY_RESULTS_FILE.write(line + '\n')
    SUMMARY_RESULTS_FILE.flush()



def close_results_files():
    ''' closes results files '''
    FULL_RESULTS_FILE.close()
    SUMMARY_RESULTS_FILE.close()



def main():
    ''' driver for validation process '''

    global search_words, included_repos, excluded_repos, excluded_endings, excluded_folders

    set_repo_mode()

    if not check_repos_json():
        create_repos_json()

    search_words = read_input_file(SEARCH_WORDS_FILENAME, critical=True)

    if REPO_MODE == INCLUDE_MODE:
        included_repos = read_repo_names(INCLUDED_REPOS_FILENAME)
    else:
        excluded_repos = read_repo_names(EXCLUDED_REPOS_FILENAME)

    excluded_endings = read_input_file(EXCLUDED_ENDINGS_FILENAME)
    excluded_folders = read_input_file(EXCLUDED_FOLDERS_FILENAME)

    read_last_updated_info()

    if not os.path.exists(OUTPUT_FOLDER):
        os.mkdir(OUTPUT_FOLDER)

    init_results_files()
    search_repos()
    close_results_files()

    write_last_updated_info()
    write_found_words()



if __name__ == "__main__":
    NOW = datetime.datetime.now().strftime("%m-%d-%Y_%H-%M-%S")
    LOGFILE = f'log_{NOW}.txt'

    if not os.path.exists(LOGS_FOLDER):
        os.mkdir(LOGS_FOLDER)

    logging.basicConfig(
        level=logging.DEBUG,
        format='%(asctime)s - LOGGING.%(levelname)s - %(message)s',
        filename=os.path.join(LOGS_FOLDER, LOGFILE),
        filemode='a'
    )
    logger = logging.getLogger(__name__)

    main()
