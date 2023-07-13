''' checks ADO repos for mentions of tables marked for deletion '''

import os
import sys
import time
import json
import logging
import re
import warnings
import datetime

import openpyxl     # xlsm, xlsx
import xlrd         # xls
import chardet
import git
from selenium import webdriver
from bs4 import BeautifulSoup


## TODO
# avoid multiple open, close for results files
# generalize wording for keywords other than table names
# consider regex options


"""
handled errors

no upstream branch
    empty repo
max retry for clone / pull
file not found
    path too long (>260)
decoding failure
"""



LOGS_FOLDER = 'logs'

REPOS_FOLDER = 'repos'
REPOS_JSON = 'repos.json'
LAST_UPDATED_FILE = 'last_updated.txt'

INPUT_FOLDER = 'input'
TABLE_NAMES_FILE = 'table_names.txt'
INCLUDED_REPOS_FILE = 'included_repos.txt'
EXCLUDED_REPOS_FILE = 'excluded_repos.txt'
EXCLUDED_ENDINGS_FILE = 'excluded_endings.txt'
EXCLUDED_DIRS_FILE = 'excluded_dirs.txt'

OUTPUT_FOLDER = 'output'
FULL_RESULTS_FILE = 'results.txt'
SUMMARY_RESULTS_FILE = 'results_summary.txt'
AFFECTED_TABLE_NAMES_FILE = 'affected.txt'

ADO_URL = 'https://dev.azure.com/bp-vsts/NAGPCCR/_apis/git/repositories?api-version=7.0'
MAX_GIT_ATTEMPTS = 5
RE_PATTERN_START = "(?<![a-z0-9_])"
RE_PATTERN_END = "(?![a-z0-9_])"
MAX_LINE_PREVIEW_LENGTH = 500
DAY_IN_SECONDS = 86400

INCLUDE_MODE, EXCLUDE_MODE = range(2)
REPO_MODE = 0

table_names = []
affected_table_names = set()
included_repos = []
excluded_repos = []
excluded_dirs = []
excluded_endings = []
last_update = {}
encodings = ['default', 'utf-8', 'utf-16']



def search_repos():
    ''' checks repos for occurrences of table names '''

    with open(REPOS_JSON, 'r', encoding="utf-8") as json_file:
        repos_json = json.loads(''.join(json_file.readlines()))

        n_repos = repos_json['count']
        for count, i in enumerate(repos_json['value']):
            repo_name = i['name'].lower().replace(' ', '%20') # no spaces in repo name
            repo_url = i['remoteUrl']
            repo_path = os.path.join(REPOS_FOLDER, repo_name)

            logger.info('repo %d/%d - %s - %s', count+1, n_repos, repo_name, repo_url)
            print(f'searching repo {count+1}/{n_repos} - {repo_name}')
            write_repo_start(repo_name, count+1)

            if ((REPO_MODE == INCLUDE_MODE) and (repo_name not in included_repos)) or \
                    ((REPO_MODE == EXCLUDE_MODE) and (repo_name in excluded_repos)):
                logger.info('excluded from validation')
                write_repo_skipped()
                write_repo_end()
                continue

            if update_repo(repo_name, repo_path, repo_url):
                write_repo_details(search_repo(repo_path))
            else:
                logger.info('skipping search')
                write_repo_skipped()

            write_repo_end()



def update_repo(name, path, url=None):
    ''' updates repo local files '''

    if os.path.exists(path):
        logger.info('repo path already exists')

        try:
            git.Repo(path).git.rev_parse('@{upstream}') # look for upstream
        except git.GitCommandError:
            logger.error('no upstream branch - empty repo?')
            return False

        if (name not in last_update) or (time.time() - last_update[name] > DAY_IN_SECONDS):
            return update_repo_core(name, path, 'pull')

        logger.info('last update less than 1 day ago - skipping pull')
        return True

    logger.info('repo path does not exist')
    return update_repo_core(name, path, 'clone', url)



def update_repo_core(name, path, mode, url=None):
    ''' main logic for repo local file update '''

    attempts = 0

    while attempts < MAX_GIT_ATTEMPTS:
        if mode == 'pull':
            if attempt_git_command(path, 'pull'):
                break
        else:
            if attempt_git_command(path, 'clone', url):
                break
        attempts += 1
        logger.info('trying again in 2 seconds')
        time.sleep(2)

    if attempts == MAX_GIT_ATTEMPTS:
        if mode == 'pull':
            logger.error('max retries for repo pull - skipping')
        else:
            logger.error('max retries for repo clone - skipping')
        return False

    last_update[name] = time.time()
    return True



def attempt_git_command(path, mode, url=None):
    ''' attempts clone or pull and returns status '''

    try:
        if mode == 'clone':
            git.Repo.clone_from(url, path)
            logger.info('clone successful')
        else:
            logger.info(git.Repo(path).git.pull())
            logger.info('pull successful')
        return True

    except git.GitCommandError as err:
        if mode == 'clone':
            logger.error('clone failed - %s', err)
        else:
            logger.error('pull failed - %s', err)
        return False



def search_repo(repo_folder):
    ''' checks repo files for occurrences of table names '''

    matches = {}
    errors = []
    searched_subfolders = []
    searched_files = []

    for root, dirs, files in os.walk(repo_folder):
        dirs[:] = [_dir for _dir in dirs if _dir.lower() not in excluded_dirs]

        # only remove if .sln file exists
        if 'packages' in dirs:
            if len([file for file in files if file.endswith('.sln')]) != 0:
                dirs.remove('packages')

        searched_subfolders += [os.path.join(root, dir) for dir in dirs]

        files[:] = [file for file in files if not file.lower().endswith(tuple(excluded_endings))]

        for file in files:
            path = os.path.join(root, file)
            logger.info(path)

            if path.endswith('.xls'):
                search_legacy_spreadsheet_file(path, matches)
                searched_files.append(path)
            elif path.endswith(tuple(['.xlsm', '.xlsx'])):
                search_spreadsheet_file(path, matches)
                searched_files.append(path)
            else:
                if search_file(path, matches):
                    searched_files.append(path)
                else:
                    errors.append(path)

    return (errors, matches, searched_subfolders, searched_files)



def search_legacy_spreadsheet_file(path, matches):
    workbook = xlrd.open_workbook(path)

    for table_name in table_names:
        pattern = RE_PATTERN_START + table_name + RE_PATTERN_END

        for sheet in workbook.sheets():
            for n_row in range(sheet.nrows):
                for n_col in range(sheet.ncols):
                    cell = sheet.cell(n_row, n_col)
                    search_cell(cell, pattern, sheet.name, n_row, n_col, matches, path, table_name)



def search_spreadsheet_file(path, matches):
    warnings.simplefilter(action='ignore', category=UserWarning)

    workbook = openpyxl.load_workbook(path, read_only=True)

    for table_name in table_names:
        pattern = RE_PATTERN_START + table_name + RE_PATTERN_END

        for sheet in workbook.worksheets:
            for n_row, row in enumerate(sheet.iter_rows()):
                for n_col, cell in enumerate(row):
                    search_cell(cell, pattern, sheet.title, n_row, n_col, matches, path, table_name)

    warnings.resetwarnings()



def search_cell(cell, pattern, sheet_name, n_row, n_col, matches, path, table_name):
    if not cell.value:
        return

    val = str(cell.value).lower()

    if re.search(pattern, val):
        if len(val) > MAX_LINE_PREVIEW_LENGTH:
            val = 'VALUE TOO LONG, LOOK AT FILE'

        log_string = f'sheet {sheet_name} row {n_row} col {n_col} - {val}'

        logger.info('match - %s - %s', table_name, log_string)
        add_new_match(matches, path, table_name, log_string)



def search_file(path, matches):
    ''' checks file for occurrences of table names '''

    decoding_result = decode_file(path)
    if not decoding_result[0]:
        return False

    lines = decoding_result[1]

    for table_name in table_names:
        pattern = RE_PATTERN_START + table_name + RE_PATTERN_END

        for line_num, line in enumerate(lines):
            line = line.lower().strip()
            if re.search(pattern, line):
                if len(line) > MAX_LINE_PREVIEW_LENGTH:
                    line = 'LINE TOO LONG, LOOK AT FILE'

                log_string = f'line {line_num} - {line}'

                logger.info('match - %s - %s', table_name, log_string)
                add_new_match(matches, path, table_name, log_string)

    return True



def add_new_match(matches, path, table_name, note):
    if not path in matches:
        matches[path] = {}

    if not table_name in matches[path]:
        matches[path][table_name] = []

    matches[path][table_name].append(note)

    affected_table_names.add(table_name)



def decode_file(path):
    ''' decodes file for reading '''

    for encoding in encodings:
        try:
            if encoding == 'default':
                with open(path, 'r') as cur_file: # intentionally didn't specify encoding
                    lines = [line.lower() for line in cur_file.readlines()]
            else:
                with open(path, 'r', encoding=encoding) as cur_file:
                    lines = [line.lower() for line in cur_file.readlines()]

            logger.info('decoding success using %s', encoding)
            return (True, lines)

        except FileNotFoundError:
            logger.error('file not found')
            return (False,)

        except (UnicodeDecodeError, UnicodeError):
            pass


    with open(path, 'rb') as binary_file:
        data = binary_file.read()

    encoding = chardet.detect(data)['encoding']

    if encoding is None:
        logger.error('decoding failure type 1')
        return (False,)

    try:
        with open(path, 'r', encoding=encoding) as cur_file:
            lines = [line.lower() for line in cur_file.readlines()]

        logger.info('decoding success using %s', encoding)
        return (True, lines)

    except (UnicodeDecodeError, UnicodeError):
        logger.error('decoding failure type 2')
        return (False,)



def check_repos_json():
    if not os.path.exists(REPOS_JSON):
        return False

    with open(REPOS_JSON, 'r', encoding="utf-8") as json_file:
        repos_json = json.loads(''.join(json_file.readlines()))

        try:
            dtm = float(repos_json['dtm'])
        except ValueError:
            logger.warning('json update dtm could not be converted to float')
            return False
        except KeyError:
            logger.error('json does not contain dtm')
            return False
        
        return time.time() - dtm < DAY_IN_SECONDS



def create_repos_json():
    ''' gets repos list using API, persists in json '''

    driver = webdriver.Chrome()
    driver.get(ADO_URL)

    input() # press enter once loaded

    body = BeautifulSoup(driver.page_source, 'html.parser').body
    driver.close()
    for pre_tag in body.find_all('pre'):
        pre_tag.unwrap()

    json_dict = json.loads(body.contents[0])
    json_dict["dtm"] = time.time()

    with open(REPOS_JSON, 'w', encoding="utf-8") as json_file:
        json.dump(json_dict, json_file, indent=4)



def read_input_file(filename, critical=False):
    path = os.path.join(INPUT_FOLDER, filename)
    if not os.path.exists(path):
        msg = f'{filename} does not exist'
        if critical:
            logger.critical(msg)
            sys.exit()
        else:
            logger.warning(msg)
            return []
    
    out = []
    with open(path, 'r', encoding="utf-8") as file:
        for line in file.readlines():
            line = line.strip().lower()
            if (not line.startswith("#")) and (not line == ""):
                out.append(line)
    return out



def read_repo_names(filename):
    
    # no spaces in repo name
    return [name.replace(' ', '%20') for name in read_input_file(filename)]



def read_last_updated_info():
    ''' reads last updated info for repos '''

    if not os.path.exists(LAST_UPDATED_FILE):
        logger.warning(f'{LAST_UPDATED_FILE} does not exist')
        return

    with open(LAST_UPDATED_FILE, 'r', encoding="utf-8") as update_file:
        for line in update_file.readlines():
            split = line.strip().split('\t')
            last_update[split[0]] = float(split[1])



def write_last_updated_info():
    ''' writes last updated info for repos '''

    with open(LAST_UPDATED_FILE, 'w', encoding="utf-8") as update_file:
        for repo, update_time in last_update.items():
            update_file.write(repo + '\t' + str(update_time) + '\n')



def init_results_files():
    ''' initializes results files '''

    with open(os.path.join(OUTPUT_FOLDER, FULL_RESULTS_FILE), 'w', encoding='utf-8') as file:
        file.write('=== CONFIG ===\n')
        file.write(f'log file - {LOGFILE}\n')
        file.write(f'repo mode - {REPO_MODE}\n')
        if REPO_MODE == INCLUDE_MODE:
            file.write(format_config_section_string('included repos', included_repos))
        else:
            file.write(format_config_section_string('excluded repos', excluded_repos))

        file.write(format_config_section_string('excluded endings', excluded_endings))
        file.write(format_config_section_string('excluded dirs', excluded_dirs))
        file.write(format_config_section_string('table names', table_names) + '\n\n')
        file.write('=== REPOS ===\n')

    with open(os.path.join(OUTPUT_FOLDER, SUMMARY_RESULTS_FILE), 'w', encoding='utf-8') as file:
        file.write('')



def format_config_section_string(st, lst):
    out = f'{st}\n'
    if len(lst) > 0:
        out += '\t' + '\n\t'.join(lst) + '\n'
    return out



def write_full_results_line(line):
    ''' writes line to full results file '''

    with open(os.path.join(OUTPUT_FOLDER, FULL_RESULTS_FILE), 'a', encoding='utf-8') as file:
        file.write(line + '\n')



def write_summary_results_line(line):
    ''' writes line to summary results file '''

    with open(os.path.join(OUTPUT_FOLDER, SUMMARY_RESULTS_FILE), 'a', encoding='utf-8') as file:
        file.write(line + '\n')



def write_repo_start(name, num):
    ''' writes repo start info to results files '''

    write_full_results_line(name + f' ({num})')
    write_summary_results_line(name + f' ({num})')



def write_repo_skipped():
    ''' writes repo skipped to results files '''

    write_full_results_line('\tskipped')
    write_summary_results_line('\tskipped')



def write_repo_details(details):
    ''' writes repo validation details to results files '''

    errors, matches, searched_subfolders, searched_files = details

    if len(matches) > 0:
        count = 0

        write_full_results_line('\tmatches:')
        for path in matches.keys():
            write_full_results_line(f'\t\t{path}')
            for table_name in matches[path].keys():
                write_full_results_line(f'\t\t\t{table_name}')
                for match in matches[path][table_name]:
                    write_full_results_line(f'\t\t\t\t{match}')
                    count += 1

        write_summary_results_line(f'\tmatches: {count}')

    if len(errors) > 0:
        write_full_results_line('\terrors:')
        for error in errors:
            write_full_results_line('\t\t' + error)

        write_summary_results_line(f'\terrors: {len(errors)}')

    if len(searched_subfolders) > 0:
        write_full_results_line('\tsubfolders searched:')
        for folder in searched_subfolders:
            write_full_results_line('\t\t' + folder)

        write_summary_results_line(f'\tsubfolders searched: {len(searched_subfolders)}')

    if len(searched_files) > 0:
        write_full_results_line('\tfiles searched:')
        for file in searched_files:
            write_full_results_line('\t\t' + file)

        write_summary_results_line(f'\tfiles searched: {len(searched_files)}')



def write_repo_end():
    ''' writes repo end to results files '''

    write_full_results_line('\n\n')
    write_summary_results_line('\n\n')



def write_affected_tables():
    sorted_affected_table_names = sorted(affected_table_names)
    with open(os.path.join(OUTPUT_FOLDER, AFFECTED_TABLE_NAMES_FILE), 'w', encoding='utf-8') as file:
        for name in sorted_affected_table_names:
            file.write(name + '\n')



def set_repo_mode():
    global REPO_MODE
    if input('enter repo mode (include - i / I, exclude - anything else): ') in ('i', 'I'):
        REPO_MODE = INCLUDE_MODE
        return
    REPO_MODE = EXCLUDE_MODE

    

def main():
    ''' driver for validation process '''
    
    global table_names, included_repos, excluded_repos, excluded_endings, excluded_dirs

    set_repo_mode()

    if not check_repos_json():
        create_repos_json()

    table_names = read_input_file(TABLE_NAMES_FILE, critical=True)

    if REPO_MODE == INCLUDE_MODE:
        included_repos = read_repo_names(INCLUDED_REPOS_FILE)
    else:
        excluded_repos = read_repo_names(EXCLUDED_REPOS_FILE)

    excluded_endings = read_input_file(EXCLUDED_ENDINGS_FILE)
    excluded_dirs = read_input_file(EXCLUDED_DIRS_FILE)

    read_last_updated_info()

    if not os.path.exists(OUTPUT_FOLDER):
        os.mkdir(OUTPUT_FOLDER)

    init_results_files()
    search_repos()
    write_last_updated_info()
    write_affected_tables()



if __name__ == "__main__":
    NOW = datetime.datetime.now().strftime("%m-%d-%Y_%H-%M-%S")
    LOGFILE = f'log_{NOW}.txt'

    if not os.path.exists(LOGS_FOLDER):
        os.mkdir(LOGS_FOLDER)

    logging.basicConfig(
        level=logging.DEBUG,
        format='%(asctime)s - LOGGING.%(levelname)s - %(message)s',
        filename=os.path.join(LOGS_FOLDER, LOGFILE),
        filemode='a'
    )
    logger = logging.getLogger(__name__)

    main()
