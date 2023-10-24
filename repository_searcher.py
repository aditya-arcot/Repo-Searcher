""" contains repository searcher class """

import os
import sys
import re
import json
import time
from datetime import datetime
from enum import Enum
from typing import Optional
import requests
import xlrd
import openpyxl
from repository import ADORepository
from logging_manager import LoggingManager
from configuration_manager import ConfigurationManager
from constants import (
    ConfigEnum,
    Constants,
    SearchTemplateModeEnum,
    SearchExcelModelEnum,
)
from results_writer import ResultsWriter, BranchSearchResults


# pylint: disable=too-few-public-methods
class RepositorySearcher:
    """used for searching repository branches for specific words"""

    search_pattern = "(?<![a-z0-9_]){word}(?![a-z0-9_])"
    max_line_preview_length = 5000
    line_too_long_msg = "VALUE TOO LONG, LOOK AT FILE"
    ado_base_url = "https://dev.azure.com/bp-vsts/NAGPCCR/_apis/git/"
    repos_url = ado_base_url + "repositories?api-version=7.0"
    branches_url = ado_base_url + "repositories/{}/refs?filter=heads/&api-version=7.0"

    def __init__(self) -> None:
        self.offline = not self.__check_internet_connection()
        date_str = datetime.now().strftime("%m-%d-%Y_%H-%M-%S")
        self.logger = LoggingManager(date_str)
        self.writer = ResultsWriter(self.logger, date_str)

        self.__init_config()
        if not self.offline:
            self.__create_repos_json()
        self.__create_repos()

        self.writer.write_config(self.config, self.search_repos_branches)

    def __check_internet_connection(self) -> bool:
        try:
            response = requests.get("https://www.google.com", timeout=5)
            return response.status_code == 200
        except requests.ConnectionError:
            return False

    def __init_config(self) -> None:
        self.config = ConfigurationManager(self.logger)
        if not self.offline:
            self.config.add_token()
        self.config.add_last_updated_info()
        self.config.add_search_words()
        self.config.add_excluded_files()
        self.config.add_excluded_folders()
        self.config.add_included_repos()
        self.config.add_excluded_repos()
        self.config.add_repo_search_template_mode()
        self.config.add_search_excel_mode()

    def __create_repos_json(self) -> None:
        """
        create json with repository information from ADO API
        learn.microsoft.com/en-us/rest/api/azure/devops/git/?view=azure-devops-rest-7.0
        """

        json_path = ConfigEnum.REPOS.value

        if os.path.exists(json_path):
            with open(json_path, "r", encoding="utf-8") as json_file:
                data = json.loads("".join(json_file.readlines()))

            if (
                "lastUpdate" in data
                and data["lastUpdate"] - time.time() < Constants.DAY_IN_SECONDS
            ):
                self.logger.info(
                    "repos json exists and was updated within the last day"
                )
                return

            self.logger.info("repos json exists but requires update")

        self.logger.info("creating repos json")

        auth = ("user", self.config.get_str(ConfigEnum.PAT.name))

        resp = self.__make_request(
            self.repos_url,
            auth,
            error_msg="max retries requesting repos, restart program",
        )

        assert resp is not None
        repos_data = resp.json()
        if "value" not in repos_data:
            self.logger.critical("badly formed json response, restart program")

        self.logger.info("success getting repos info")

        branch_start = "refs/heads/"

        for pos, repo in enumerate(repos_data["value"]):
            if "defaultBranch" not in repo:
                self.logger.error(
                    "repo does not have a default branch - most likely empty"
                )
                repos_data["value"][pos]["branches"] = []
                continue

            repo["defaultBranch"] = repo["defaultBranch"].replace(branch_start, "", 1)

            self.logger.info(
                f'getting branch info for {repo["name"]} ({pos+1}/{repos_data["count"]})'
            )

            url = self.branches_url.format(repo["id"])
            resp = self.__make_request(
                url, auth, error_msg="max retries requesting branches, skipping"
            )
            if not resp:
                continue

            branches_data = resp.json()
            branches = [
                branch["name"].replace(branch_start, "", 1)
                for branch in branches_data["value"]
            ]
            assert repo["defaultBranch"] in branches
            repos_data["value"][pos]["branches"] = branches

        repos_data["lastUpdate"] = time.time()

        try:
            with open(json_path, "w", encoding="utf-8") as json_file:
                json.dump(repos_data, json_file, indent=4)
        except requests.exceptions.JSONDecodeError:
            self.logger.critical("error parsing repos json info")

    def __make_request(self, url, auth, max_request_attempts=5, error_msg=""):
        attempts = 0
        resp = None
        req_err = "request failed, trying again"

        while attempts < max_request_attempts:
            try:
                resp = requests.get(url, auth=auth, timeout=10)
                if resp.status_code == 200:
                    return resp
                attempts += 1
                self.logger.error(req_err)

            except requests.exceptions.ConnectionError:
                attempts += 1
                self.logger.error(req_err)

        self.logger.critical(error_msg)
        sys.exit()

    def __create_repos(self) -> None:
        json_path = ConfigEnum.REPOS.value

        if not os.path.exists(json_path):
            self.logger.critical("repos json does not exist")

        with open(json_path, "r", encoding="utf-8") as json_file:
            data = json.loads("".join(json_file.readlines()))
        for repo_info in data["value"]:
            repo_info["name"] = repo_info["name"].replace(" ", Constants.ENCODED_SPACE)

        self.search_repos_branches = self.__create_search_repos_info(data["value"])

        self.repos: list[ADORepository] = []
        for repo, branches in self.search_repos_branches.items():
            url = next(
                (d["remoteUrl"] for d in data["value"] if d.get("name") == repo), None
            )
            if url is None:
                self.logger.error(f"{repo} url not found")
                continue
            self.repos.append(
                ADORepository(
                    self.logger, repo, branches, url, os.path.join("repos", repo)
                )
            )

    class RepoDetailsListEnum(Enum):
        """enumerations for repo details list"""

        DEFAULT_BRANCH, BRANCHES = range(2)

    def __create_search_repos_info(self, repo_json):
        repo_details = {
            repo["name"]: [
                repo["defaultBranch"] if "defaultBranch" in repo else "",
                repo["branches"],
            ]
            for repo in repo_json
        }

        match self.config.get_int(ConfigEnum.SEARCH_TEMPLATE_MODE.name):
            case SearchTemplateModeEnum.ALL_REPOS_DEFAULT_BRANCH.value:
                search_repos_branches = {
                    name: set([details[self.RepoDetailsListEnum.DEFAULT_BRANCH.value]])
                    for name, details in repo_details.items()
                    if len(details[self.RepoDetailsListEnum.BRANCHES.value]) != 0
                }
            case SearchTemplateModeEnum.ALL_REPOS_ALL_BRANCHES.value:
                search_repos_branches = {
                    name: set(details[self.RepoDetailsListEnum.BRANCHES.value])
                    for name, details in repo_details.items()
                }
            case _:  # default
                search_repos_branches = {}

        self.__add_included_branches(repo_details, search_repos_branches)
        self.__remove_excluded_branches(repo_details)

        return search_repos_branches

    def __add_included_branches(self, repo_details: dict, search_repos_branches: dict):
        included = self.config.get_dict(ConfigEnum.INCLUDED_REPOS.name)
        for include_name, include_branches in included.items():
            if include_name not in repo_details:
                self.logger.error(f"{include_name} is not a valid repo")
                continue

            default = repo_details[include_name][
                self.RepoDetailsListEnum.DEFAULT_BRANCH.value
            ]
            branches = repo_details[include_name][
                self.RepoDetailsListEnum.BRANCHES.value
            ]

            if include_name not in search_repos_branches:
                search_repos_branches[include_name] = set()

            # add default branch
            if len(include_branches) == 0 and default:
                search_repos_branches[include_name].add(default)

            # add all branches
            elif "*" in include_branches:
                search_repos_branches[include_name].update(branches)

            # add specific branches
            else:
                for include_branch in include_branches:
                    if include_branch in branches:
                        search_repos_branches[include_name].add(include_branch)
                    else:
                        self.logger.error(
                            f"{include_branch} is not a branch of {include_name}"
                        )

    def __remove_excluded_branches(self, search_repos_branches: dict):
        excluded = self.config.get_dict(ConfigEnum.EXCLUDED_REPOS.name)
        for exclude_name, exclude_branches in excluded.items():
            if exclude_name not in search_repos_branches:
                self.logger.info(
                    f"{exclude_name} for {exclude_name} is already not included in search"
                )
                continue

            # remove all branches
            if len(exclude_branches) == 0 or "*" in exclude_branches:
                search_repos_branches.pop(exclude_name)

            else:
                for exclude_branch in exclude_branches:
                    if exclude_branch in search_repos_branches[exclude_name]:
                        search_repos_branches[exclude_name].remove(exclude_branch)
                    else:
                        self.logger.info(
                            f"branch {exclude_branch} is already not included in search"
                        )

    def search(self) -> None:
        """search through repositories for specified words"""
        last_update_dict = self.config.get_dict(ConfigEnum.LAST_UPDATE.name)
        no_search = len(self.config.get_list(ConfigEnum.SEARCH_WORDS.name)) == 0

        self.logger.info(f"offline mode: {self.offline}")

        for repo_ind, repo in enumerate(self.repos):
            self.logger.info(f"repo {repo_ind+1}/{len(self.repos)} - {repo.name}")
            self.writer.write_repo_start(repo.name)

            for branch_ind, branch in enumerate(repo.branches):
                self.logger.info(
                    f"branch {branch_ind+1}/{len(repo.branches)} - {branch}"
                )
                self.writer.write_branch_start(branch)

                if not self.offline:
                    if repo.update_branch(branch, last_update_dict):
                        self.__search_branch_main(no_search, repo, branch)
                    else:
                        self.__skip_branch("update unsuccessful")

                else:
                    try:
                        self.logger.info(last_update_dict[repo.name][branch])
                    except KeyError:
                        self.logger.error("branch last update info not available")

                    self.__search_branch_main(no_search, repo, branch)

        if not no_search:
            self.writer.write_found_words()
        self.__write_last_update()

    def __search_branch_main(
        self, no_search: bool, repo: ADORepository, branch: str
    ) -> None:
        if no_search:
            self.__skip_branch("no search attempted")
        else:
            details = self.__search_branch(repo, branch)
            if details:
                self.writer.write_branch_results(repo.name, branch, details)

    def __skip_branch(self, msg: str) -> None:
        """log, write info for skipped branch"""
        self.logger.info(msg)
        self.writer.write_branch_skip(msg)

    def __search_branch(
        self, repo: ADORepository, branch: str
    ) -> Optional[BranchSearchResults]:
        """search local repository files for search words"""
        branch_path = os.path.join(repo.path, branch)
        if not os.path.exists(branch_path):
            self.logger.error("branch path does not exist")
            return None

        search_results = BranchSearchResults()
        excluded_folders = self.config.get_list(ConfigEnum.EXCLUDED_FOLDERS.name)
        excluded_files = self.config.get_list(ConfigEnum.EXCLUDED_FILES.name)

        for root, dirs, files in os.walk(branch_path):
            skipped_dirs = [_dir for _dir in dirs if _dir.lower() in excluded_folders]
            skipped_files = [
                file for file in files if file.lower().endswith(tuple(excluded_files))
            ]

            search_results.skipped_folders += [
                os.path.join(root, _dir) + os.sep for _dir in skipped_dirs
            ]
            search_results.skipped_files += [
                os.path.join(root, file) for file in skipped_files
            ]

            # modify in place to prune search tree
            dirs[:] = [_dir for _dir in dirs if _dir not in skipped_dirs]
            files[:] = [file for file in files if file not in skipped_files]

            search_results.folders += [os.path.join(root, dir) for dir in dirs]

            for file in files:
                path = os.path.join(root, file)
                self.logger.info(f"file - {path}", stdout=False)

                if path.endswith(tuple([".xls", ".xlsm", ".xlsx"])):
                    if (
                        self.config.get_int(ConfigEnum.EXCEL_SEARCH_MODE.name)
                        == SearchExcelModelEnum.NO.value
                    ):
                        search_results.skipped_files.append(path)
                    else:
                        search_results.files.append(path)
                        if path.endswith(".xls"):
                            self.__search_legacy_spreadsheet_file(
                                path, search_results.matches
                            )
                        else:
                            self.__search_spreadsheet_file(path, search_results.matches)
                else:
                    self.__search_plaintext_file(path, search_results)

        return search_results

    def __search_legacy_spreadsheet_file(self, path: str, matches: dict) -> None:
        """searches .xls file by iterating over words, sheets, cells"""
        workbook = xlrd.open_workbook(path)
        for search_word in self.config.get_list(ConfigEnum.SEARCH_WORDS.name):
            pattern = self.search_pattern.format(word=search_word)
            for sheet in workbook.sheets():
                for n_row in range(sheet.nrows):
                    for n_col in range(sheet.ncols):
                        cell = sheet.cell(n_row, n_col)
                        self.__search_cell(
                            cell,
                            sheet.name,
                            n_row,
                            n_col,
                            matches,
                            path,
                            pattern,
                            search_word,
                        )

    def __search_spreadsheet_file(self, path: str, matches: dict) -> None:
        """searches .xlsm or .xlsx file by iterating over words, sheets, cells"""
        workbook = openpyxl.load_workbook(path, read_only=True)
        for search_word in self.config.get_list(ConfigEnum.SEARCH_WORDS.name):
            pattern = self.search_pattern.format(word=search_word)
            for sheet in workbook.worksheets:
                for n_row, row in enumerate(sheet.iter_rows()):
                    for n_col, cell in enumerate(row):
                        self.__search_cell(
                            cell,
                            sheet.title,
                            n_row,
                            n_col,
                            matches,
                            path,
                            pattern,
                            search_word,
                        )

    # pylint: disable=too-many-arguments
    def __search_cell(
        self,
        cell,
        sheet_name: str,
        n_row: int,
        n_col: int,
        matches: dict,
        path: str,
        pattern: str,
        search_word: str,
    ) -> None:
        """searches cell for search pattern"""
        if not cell.value:
            return

        val = str(cell.value).lower()

        if re.search(pattern, val):
            if len(val) > self.max_line_preview_length:
                val = self.line_too_long_msg
            log_string = f"sheet {sheet_name}, row {n_row}, col {n_col} - {val}"
            self.logger.info(f"match - {search_word} - {log_string}", stdout=False)
            self.__add_new_match(matches, path, search_word, log_string)

    def __search_plaintext_file(
        self, path: str, search_results: BranchSearchResults
    ) -> None:
        """searches non-spreadsheet file for specified words"""
        lines = self.__decode_file(path, search_results)
        if len(lines) == 0:
            return

        for search_word in self.config.get_list(ConfigEnum.SEARCH_WORDS.name):
            pattern = self.search_pattern.format(word=search_word)
            for line_num, line in enumerate(lines):
                line = line.lower().strip()

                if re.search(pattern, line):
                    if len(line) > self.max_line_preview_length:
                        line = self.line_too_long_msg
                    log_string = f"line {line_num + 1} - {line}"
                    self.logger.info(
                        f"match - {search_word} - {log_string}", stdout=False
                    )
                    self.__add_new_match(
                        search_results.matches, path, search_word, log_string
                    )

    def __decode_file(self, path: str, search_results: BranchSearchResults) -> list:
        """attempts decoding file, returns lines"""
        try:
            with open(path, "r", encoding="utf-8", errors="ignore") as cur_file:
                lines = [line.lower() for line in cur_file.readlines()]
            self.logger.info("decoding success", stdout=False)
            search_results.files.append(path)
            return lines

        except FileNotFoundError:
            self.logger.error(f"file not found - path too long - {path}", stdout=False)

        except (UnicodeDecodeError, UnicodeError):
            self.logger.error("decoding failure", stdout=False)

        search_results.errors.append(path)
        return []

    def __add_new_match(
        self, matches: dict, path: str, search_word: str, note: str
    ) -> None:
        """adds new match to matches dict"""
        if not path in matches:
            matches[path] = {}

        if not search_word in matches[path]:
            matches[path][search_word] = []

        matches[path][search_word].append(note)

    def __write_last_update(self) -> None:
        """writes last updated info to file"""
        update_dict = self.config.get_dict(ConfigEnum.LAST_UPDATE.name)
        with open(ConfigEnum.LAST_UPDATE.value, "w", encoding="utf-8") as update_file:
            json.dump(update_dict, update_file, indent=4)
